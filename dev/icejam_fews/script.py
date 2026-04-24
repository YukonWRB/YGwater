import openpyxl
from openpyxl.utils import get_column_letter

try:
    wb = openpyxl.load_workbook(r'C:\Users\esniede\Documents\Breakup_model_YukonRiver_2025.xlsx', data_only=False)

    print("=== SHEET NAMES ===")
    print(wb.sheetnames)

    print("\n=== NAMED RANGES ===")
    for name in wb.defined_names.definedName:
        print(f"  {name.name}: {name.attr_text}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n\n{'='*60}")
        print(f"SHEET: {sheet_name}")
        print(f"Dimensions: {ws.dimensions}")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        print("\n--- CELL VALUES AND FORMULAS ---")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    col_letter = get_column_letter(cell.column)
                    print(f"  {col_letter}{cell.row}: {repr(cell.value)}")
        
        print("\n--- MERGED CELLS ---")
        for merge in ws.merged_cells.ranges:
            print(f"  {merge}")
        
        print("\n--- CHARTS ---")
        for chart in ws._charts:
            print(f"  Chart type: {type(chart).__name__}")
            print(f"  Title: {chart.title}")
            try:
                for series in chart.series:
                    print(f"    Series title: {series.title}")
                    print(f"    Series values: {series.val}")
                    print(f"    Series x-values: {series.xVal if hasattr(series, 'xVal') else 'N/A'}")
            except Exception as e:
                print(f"  Error reading series: {e}")
        
        print("\n--- DATA VALIDATIONS ---")
        for dv in ws.data_validations.dataValidation:
            print(f"  Range: {dv.sqref}, Type: {dv.type}, Formula1: {dv.formula1}, Formula2: {dv.formula2}")
except Exception as e:
    print(f"An error occurred: {e}")
